import openpyxl

book = openpyxl.load_workbook('REST/test.xlsx')

sheet = book.active

a1 = sheet['A1']
a2 = sheet['A2']
a3 = sheet.cell(row=3, column=1)

print(a1.value)
print(a2.value) 
print(a3.value)


sheet1 = book.create_sheet('Sheet2')


book.save('REST/test.xlsx')